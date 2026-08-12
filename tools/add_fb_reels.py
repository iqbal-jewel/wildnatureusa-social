"""Add one Facebook fact Reel a day to the content plan, at 19:00 ET.

19:00 sits directly between the two strongest Facebook slots measured so far
(18:00 quiz and 20:00 fact -- see docs/reels.md), doesn't collide with any
existing Facebook time, and is already proven as a good evening slot on the
Instagram side (its 19:00 ET quiz).

The facts are not invented, same principle as tools/add_ig_facts.py: every
reel reuses a real fact already in the plan, taken from the Facebook day 45
days away (a different offset than the 92 IG already uses, so the two new
formats aren't quietly drawing on the same source day) using that day's
20:00 ET fact -- the last slot of the day, so the day's "headline" fact
becomes that day's reel too.

Idempotent: existing FB-RL rows are left alone and the run is a no-op.

    python tools/add_fb_reels.py            # dry run, reports what it would add
    python tools/add_fb_reels.py --write
"""
import argparse
import collections
import datetime as dt
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PLAN = ROOT / "plan" / "content_plan_fixed.xlsx"

C_ID, C_PLATFORM, C_TYPE, C_DATE, C_DOW = 0, 1, 2, 3, 4
C_TIME, C_ANIMAL, C_CATEGORY, C_CAPTION = 5, 6, 7, 8
C_HASHTAGS, C_STATUS = 17, 18

POST_TYPE = "Reel Post (Fact)"
SLOT = "7:00 PM"          # 19:00 ET -- between the 18:00 quiz and 20:00 fact
SOURCE_PICK = 3            # index into a day's 4 FB facts: 08,12,16,[20]
DAY_OFFSET = 45            # different from IG's 92, so the two rarely overlap


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true", help="save the workbook")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(PLAN)
    ws = wb["Content Calendar"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2) if r[C_ID].value]

    if any(str(r[C_ID]).startswith("FB-RL-") for r in rows):
        print("FB-RL rows already present -- nothing to do")
        return 0

    fb_fact = [r for r in rows
              if r[C_PLATFORM] == "Facebook" and r[C_TYPE] == "Text Post (Fact)"]
    by_date = collections.defaultdict(list)
    for r in rows:
        by_date[str(r[C_DATE])[:10]].append(r)
    fact_by_date = collections.defaultdict(list)
    for r in fb_fact:
        fact_by_date[str(r[C_DATE])[:10]].append(r)
    for d in fact_by_date:
        fact_by_date[d].sort(key=lambda r: dt.datetime.strptime(
            str(r[C_TIME]).strip().upper(), "%I:%M %p").time())

    dates = sorted(fact_by_date)
    print(f"{len(fb_fact)} Facebook facts across {len(dates)} days")

    short = [d for d in dates if len(fact_by_date[d]) <= SOURCE_PICK]
    if short:
        print(f"ERROR: {len(short)} day(s) have too few facts to draw from, "
              f"e.g. {short[:3]}")
        return 1

    new_rows = []
    for i, date in enumerate(dates):
        src_date = dates[(i + DAY_OFFSET) % len(dates)]
        src = fact_by_date[src_date][SOURCE_PICK]
        row = [None] * ws.max_column
        row[C_ID] = f"FB-RL-{i + 1:04d}"
        row[C_PLATFORM] = "Facebook"
        row[C_TYPE] = POST_TYPE
        row[C_DATE] = date
        row[C_DOW] = dt.date.fromisoformat(date).strftime("%A")
        row[C_TIME] = SLOT
        row[C_ANIMAL] = src[C_ANIMAL]
        row[C_CATEGORY] = src[C_CATEGORY]
        row[C_CAPTION] = src[C_CAPTION]
        row[C_HASHTAGS] = src[C_HASHTAGS]
        row[C_STATUS] = "Not Posted"
        new_rows.append(row)

    # No reel may share a date with any existing post carrying the same text,
    # on either platform.
    clash = [(r[C_DATE], r[C_CAPTION]) for r in new_rows
             if any(f[C_CAPTION] == r[C_CAPTION] for f in by_date[r[C_DATE]])]
    print(f"same-day clashes with existing content: {len(clash)}")
    for d, cap in clash[:5]:
        print(f"  {d}: {str(cap)[:70]}")

    animals = {r[C_ANIMAL] for r in new_rows}
    print(f"\nadding {len(new_rows)} Facebook reels, {len(animals)} species")
    print(f"slot: {SLOT} ET, one per day")
    for r in new_rows[:3]:
        print(f"  {r[C_ID]} {r[C_DATE]} {r[C_TIME]:>8}  {r[C_ANIMAL]}")
        print(f"      {str(r[C_CAPTION])[:78]}")

    if not args.write:
        print("\nDRY RUN -- pass --write to save")
        return 1 if clash else 0

    if clash:
        print("\nERROR: refusing to write with same-day clashes present")
        return 1

    for row in new_rows:
        ws.append(row)
    wb.save(PLAN)
    print(f"\nwrote {PLAN} -- now {ws.max_row - 1} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
