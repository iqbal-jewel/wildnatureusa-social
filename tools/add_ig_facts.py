"""Add two Instagram fact posts a day to the content plan.

Instagram currently gets only the two daily quiz images; the fact content runs
on Facebook alone, which is backwards given Instagram carries the far larger
audience.

The facts themselves are *not* invented. This is a wildlife facts page, and
making up zoology to fill slots would be worse than posting nothing. Every IG
fact reuses a real fact already in the plan, sourced from the original
workbook -- taken from the Facebook day 92 days away (half the run), so the
same sentence never lands on both platforms near the same date.

Idempotent: existing IG-FCT rows are left alone and the run is a no-op.

    python tools/add_ig_facts.py            # dry run, reports what it would add
    python tools/add_ig_facts.py --write
"""
import argparse
import collections
import datetime as dt
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PLAN = ROOT / "plan" / "content_plan_fixed.xlsx"

C_ID, C_PLATFORM, C_TYPE, C_DATE, C_DOW = 0, 1, 2, 3, 4
C_TIME, C_ANIMAL, C_CATEGORY, C_CAPTION = 5, 6, 7, 8
C_HASHTAGS, C_STATUS = 17, 18

POST_TYPE = "Image Post (Fact)"

# The quiz slots are 11:00 and 19:00 ET; these sit clear of both and of each
# other, in the two windows the plan is not already using.
IG_SLOTS = ["8:00 AM", "3:00 PM"]

# Which of each source day's four Facebook facts to lift (08:00 and 16:00 ET).
SOURCE_PICKS = (0, 2)

# Half the 184-day run: guarantees maximum separation between the same fact
# appearing on Facebook and on Instagram.
DAY_OFFSET = 92


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true", help="save the workbook")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(PLAN)
    ws = wb["Content Calendar"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2) if r[C_ID].value]

    if any(str(r[C_ID]).startswith("IG-FCT-") for r in rows):
        print("IG-FCT rows already present -- nothing to do")
        return 0

    # Facebook facts, grouped by date and ordered within the day.
    fb = [r for r in rows
          if r[C_PLATFORM] == "Facebook" and r[C_TYPE] == "Text Post (Fact)"]
    by_date = collections.defaultdict(list)
    for r in fb:
        by_date[str(r[C_DATE])[:10]].append(r)
    for d in by_date:
        by_date[d].sort(key=lambda r: dt.datetime.strptime(
            str(r[C_TIME]).strip().upper(), "%I:%M %p").time())

    dates = sorted(by_date)
    print(f"{len(fb)} Facebook facts across {len(dates)} days")

    short = [d for d in dates if len(by_date[d]) <= max(SOURCE_PICKS)]
    if short:
        print(f"ERROR: {len(short)} day(s) have too few facts to draw from, "
              f"e.g. {short[:3]}")
        return 1

    new_rows, seen_pairs = [], set()
    n = 0
    for i, date in enumerate(dates):
        src_date = dates[(i + DAY_OFFSET) % len(dates)]
        for slot, pick in zip(IG_SLOTS, SOURCE_PICKS):
            src = by_date[src_date][pick]
            n += 1
            row = [None] * ws.max_column
            row[C_ID] = f"IG-FCT-{n:04d}"
            row[C_PLATFORM] = "Instagram"
            row[C_TYPE] = POST_TYPE
            row[C_DATE] = date
            row[C_DOW] = dt.date.fromisoformat(date).strftime("%A")
            row[C_TIME] = slot
            row[C_ANIMAL] = src[C_ANIMAL]
            row[C_CATEGORY] = src[C_CATEGORY]
            row[C_CAPTION] = src[C_CAPTION]
            row[C_HASHTAGS] = src[C_HASHTAGS]
            row[C_STATUS] = "Not Posted"
            new_rows.append(row)
            seen_pairs.add((date, src[C_CAPTION]))

    # No IG fact may share a date with the Facebook post carrying the same text.
    clash = [(r[C_DATE], r[C_CAPTION]) for r in new_rows
             if any(f[C_CAPTION] == r[C_CAPTION] and str(f[C_DATE])[:10] == r[C_DATE]
                    for f in by_date[r[C_DATE]])]
    print(f"same-day clashes with the Facebook copy: {len(clash)}")

    animals = {r[C_ANIMAL] for r in new_rows}
    print(f"adding {len(new_rows)} Instagram fact posts, {len(animals)} species")
    print(f"slots: {', '.join(IG_SLOTS)} ET")
    for r in new_rows[:3]:
        print(f"  {r[C_ID]} {r[C_DATE]} {r[C_TIME]:>8}  {r[C_ANIMAL]}")
        print(f"      {str(r[C_CAPTION])[:78]}")

    if not args.write:
        print("\nDRY RUN -- pass --write to save")
        return 0

    for row in new_rows:
        ws.append(row)
    wb.save(PLAN)
    print(f"\nwrote {PLAN} -- now {ws.max_row - 1} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
